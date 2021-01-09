from django.views.generic.list import ListView
from django.views.generic.detail import DetailView
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.urls import reverse, reverse_lazy
from .models import Page
from .forms import PageForm
from django.shortcuts import render, redirect
from django.conf import settings
from io import BytesIO
from reportlab.pdfgen import canvas
from django.views.generic import View
from django.http import HttpResponse
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from django.utils.html import strip_tags

from django.contrib.admin.views.decorators import staff_member_required
from django.utils.decorators import method_decorator

#Vista genérica para mostrar resultados
from django.views.generic.base import TemplateView
from django.views.generic import TemplateView
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
#Nos devuelve un objeto resultado, en este caso un archivo de excel
from django.http.response import HttpResponse

#mixin
class StaffRequiredMixin(object):
    """mixin requerira que el usuario sea miembro del staff
    """
    @method_decorator(staff_member_required)
    def dispatch(self, request, *args, **kwargs):
        # if not request.user.is_staff:
        #     return redirect(reverse_lazy('admin:login'))
        return super(StaffRequiredMixin, self).dispatch(request, *args, **kwargs)

# Create your views here.
class PageListView(ListView):
    model = Page

class PageDetailView(DetailView):
    model = Page

@method_decorator(staff_member_required, name='dispatch')
class PageCreate(CreateView):
    model = Page
    form_class = PageForm
    success_url = reverse_lazy('pages:pages')

@method_decorator(staff_member_required, name='dispatch')
class PageUpdate(UpdateView):
    model = Page
    form_class = PageForm
    template_name_suffix = '_update_form'
    
    def get_success_url(self):
        return reverse_lazy('pages:update', args=[self.object.id]) + '?ok'

@method_decorator(staff_member_required, name='dispatch')
class PageDelete(DeleteView):
    model = Page
    success_url = reverse_lazy('pages:pages')

class BuscarView(TemplateView):
    def post(self, request, *args, **kwargs):
        buscar=request.POST['buscalo']
        pages = Page.objects.filter(title__contains=buscar)
        print(pages)
        return render(request, 'pages/buscar.html',
                    {'pages':pages, 'page':True})

#formato PDF
class ReportePDF(View):  
     
    #def cabecera(self,pdf):
        #Utilizamos el archivo logo_django.png que está guardado en la carpeta media/imagenes
        #archivo_imagen = settings.MEDIA_ROOT+'/imagenes/logo_django.png'
        #Definimos el tamaño de la imagen a cargar y las coordenadas correspondientes
        #pdf.drawImage(archivo_imagen, 40, 750, 120, 90,preserveAspectRatio=True)

    def tabla(self,pdf,y):
        #Creamos una tupla de encabezados para neustra tabla
        encabezados = ('ID', 'Producto', 'Detalle', 'UND', 'KG')
        #Creamos una lista de tuplas que van a contener a las personas
        detalles = [(page.id, page.title, strip_tags(page.content), page.order, page.cimal) for page in Page.objects.all()]
        #Establecemos el tamaño de cada una de las columnas de la tabla
        detalle_orden = Table([encabezados] + detalles, colWidths=[1.5 * cm, 5 * cm, 6 * cm, 2 * cm, 2 * cm,])
        #Aplicamos estilos a las celdas de la tabla
        detalle_orden.setStyle(TableStyle(
            [
                #La primera fila(encabezados) va a estar centrada
                ('ALIGN',(0,0),(3,0),'CENTER'),
                #Los bordes de todas las celdas serán de color negro y con un grosor de 1
                ('GRID', (0, 0), (-1, -1), 1, colors.black), 
                #El tamaño de las letras de cada una de las celdas será de 10
                ('FONTSIZE', (0, 0), (-1, -1), 10),
            ]
        ))
        #Establecemos el tamaño de la hoja que ocupará la tabla 
        detalle_orden.wrapOn(pdf, 800, 600)
        #Definimos la coordenada donde se dibujará la tabla
        detalle_orden.drawOn(pdf, 60,y)              
         
    def get(self, request, *args, **kwargs):
        #Indicamos el tipo de contenido a devolver, en este caso un pdf
        response = HttpResponse(content_type='application/pdf')
        #La clase io.BytesIO permite tratar un array de bytes como un fichero binario, se utiliza como almacenamiento temporal
        buffer = BytesIO()
        #Canvas nos permite hacer el reporte con coordenadas X y Y
        pdf = canvas.Canvas(buffer)
        #Llamo al método cabecera donde están definidos los datos que aparecen en la cabecera del reporte.
        #self.cabecera(pdf)
        #Establecemos el tamaño de letra en 16 y el tipo de letra Helvetica
        pdf.setFont("Helvetica", 16)
        #Dibujamos una cadena en la ubicación X,Y especificada
        pdf.drawString(230, 790, u"Control Inventario")
        pdf.setFont("Helvetica", 14)
        #pdf.drawString(200, 770, u"REPORTE DE PERSONAS")
        y = 600
        self.tabla(pdf, y)
        #Con show page hacemos un corte de página para pasar a la siguiente
        pdf.showPage()
        pdf.save()
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response

#fin formato PDF

#formato Excel

#Nuestra clase hereda de la vista genérica TemplateView
class ReporteExcel(TemplateView):
     
    #Usamos el método get para generar el archivo excel 
    def get(self, request, *args, **kwargs):
        #Obtenemos todas las pages de nuestra base de datos
        pages = Page.objects.all()
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws['B1'] = 'Inventario'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = 'ID'
        ws['C3'] = 'PRODUCTO'
        ws['D3'] = 'DETALLE'
        ws['E3'] = 'UND'
        ws['F3'] = 'KG'       
        cont=5
        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for page in pages:
            ws.cell(row=cont,column=2).value = page.id
            ws.cell(row=cont,column=3).value = page.title
            ws.cell(row=cont,column=4).value = strip_tags(page.content)
            ws.cell(row=cont,column=5).value = page.order
            ws.cell(row=cont,column=6).value = page.cimal
            cont = cont + 1
        #Establecemos el nombre del archivo
        nombre_archivo ="ReporteExcel.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

#fin formato Excel

   


